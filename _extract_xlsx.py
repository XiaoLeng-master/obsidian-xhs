#!/usr/bin/env python3
"""Extract all data from XiaoHongShu xlsx file."""
import openpyxl

wb = openpyxl.load_workbook('/Users/lvran/ObsidianVaults_2/小红书商品抓取.xlsx', data_only=True)
print("Sheet names:", wb.sheetnames)
for name in wb.sheetnames:
    ws = wb[name]
    print(f"\n===== Sheet: {name} (rows={ws.max_row}, cols={ws.max_column}) =====")
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 1000), values_only=False), 1):
        vals = []
        for cell in row:
            vals.append(str(cell.value) if cell.value is not None else "")
        line = " | ".join(vals)
        print(f"[R{row_idx}] {line}")
