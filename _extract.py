#!/usr/bin/env python3
"""Extract text from both xlsx and PDF files."""
import sys
import os

os.chdir('/Users/lvran/ObsidianVaults_2')

# ===== Extract xlsx =====
print("=" * 60)
print("EXTRACTING XLSX FILE")
print("=" * 60)

# Try openpyxl first
try:
    import openpyxl
    wb = openpyxl.load_workbook('小红书商品抓取.xlsx', data_only=True)
    print(f"Sheet names: {wb.sheetnames}")
    for name in wb.sheetnames:
        ws = wb[name]
        print(f"\n===== Sheet: {name} (rows={ws.max_row}, cols={ws.max_column}) =====")
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 2000), values_only=False), 1):
            vals = []
            for cell in row:
                v = cell.value
                vals.append(str(v) if v is not None else "")
            line = " | ".join(vals)
            print(f"[R{row_idx}] {line}")
    print("\n[XLSX extraction completed with openpyxl]")
except ImportError:
    print("[openpyxl not available, trying zipfile + xml approach]")
    # Manual xlsx parsing via zipfile + xml
    import zipfile
    import xml.etree.ElementTree as ET

    with zipfile.ZipFile('小红书商品抓取.xlsx', 'r') as z:
        # Read shared strings
        shared_strings = []
        try:
            with z.open('xl/sharedStrings.xml') as f:
                tree = ET.parse(f)
                ns = {'s': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                for si in tree.findall('.//s:si', ns):
                    text_parts = []
                    for t in si.iter('{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t'):
                        if t.text:
                            text_parts.append(t.text)
                    shared_strings.append(''.join(text_parts))
            print(f"Read {len(shared_strings)} shared strings")
        except KeyError:
            print("No shared strings found")

        # Read workbook to get sheet names
        with z.open('xl/workbook.xml') as f:
            tree = ET.parse(f)
            ns = {'s': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
            sheets = tree.findall('.//s:sheet', ns)
            sheet_info = [(s.get('name'), s.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')) for s in sheets]

        # Read each sheet
        for idx, (sheet_name, _) in enumerate(sheet_info, 1):
            sheet_path = f'xl/worksheets/sheet{idx}.xml'
            print(f"\n===== Sheet: {sheet_name} =====")
            try:
                with z.open(sheet_path) as f:
                    tree = ET.parse(f)
                    ns = {'s': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                    rows = tree.findall('.//s:row', ns)
                    for row in rows:
                        row_num = row.get('r')
                        cells = row.findall('s:c', ns)
                        vals = []
                        for cell in cells:
                            cell_type = cell.get('t')
                            v_el = cell.find('s:v', ns)
                            value = v_el.text if v_el is not None else ""
                            if cell_type == 's' and value and shared_strings:
                                try:
                                    value = shared_strings[int(value)]
                                except (IndexError, ValueError):
                                    pass
                            vals.append(value if value else "")
                        line = " | ".join(vals)
                        print(f"[R{row_num}] {line}")
            except KeyError:
                print(f"Sheet file not found: {sheet_path}")
    print("\n[XLSX extraction completed with zipfile+xml]")

print("\n\n")
print("=" * 60)
print("EXTRACTING PDF FILE")
print("=" * 60)

# Try extracting PDF
try:
    import pdfplumber
    with pdfplumber.open('虚拟商品机会分析报告.pdf') as pdf:
        print(f"Total pages: {len(pdf.pages)}")
        for i, page in enumerate(pdf.pages):
            text = page.extract_text()
            print(f"\n===== Page {i+1} =====")
            if text:
                print(text)
            else:
                print("[No text extracted from this page - might be image-based]")
    print("\n[PDF extraction completed with pdfplumber]")
except ImportError:
    print("[pdfplumber not available, trying PyPDF2]")
    try:
        import PyPDF2
        with open('虚拟商品机会分析报告.pdf', 'rb') as f:
            reader = PyPDF2.PdfReader(f)
            print(f"Total pages: {len(reader.pages)}")
            for i, page in enumerate(reader.pages):
                text = page.extract_text()
                print(f"\n===== Page {i+1} =====")
                if text:
                    print(text)
                else:
                    print("[No text extracted]")
        print("\n[PDF extraction completed with PyPDF2]")
    except ImportError:
        print("[Neither pdfplumber nor PyPDF2 available. Trying to install...]")
